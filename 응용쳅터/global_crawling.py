import requests
import json
import openpyxl
from bs4 import BeautifulSoup as bs

wb = openpyxl.Workbook()

ws = wb.create_sheet('오즈키즈, 농심 크롤링')

ws['A1'] = 'product_price'
ws['C1'] = 'product_code'
ws['E1'] = 'image_url'

# func1이 실행되면 shop의 json을 가진 사이트를 파싱
def func1(list, k, app_key) :
    try:
        requestdata = requests.get(f"https://{list}.cafe24api.com/api/v2/products/{k}?shop_no=1&cafe24_app_key={app_key}&cafe24_api_version=2022-06-01")
        return requestdata
    except:
        return func2(list, k, app_key)      


# func2가 실행되면 cafe의 json을 가진 사이트를 파싱
def func2(list, k, app_key) : 
    try:
        requestdata = requests.get(f"https://{list}.cafe24api.com/api/v2/products/{k}?cafe24_app_key={app_key}")
        
        return requestdata
    except:
        return func3()

# 모든 function이 실행되지 않으면 어쩔 수 없이 html 전체를 파싱해서 값을 가져오기(ex 메디큐브)
def func3() :
    page = requests.get("https://themedicube.co.kr/product/detail.html?product_no=1103&cate_no=466&display_group=2")
    soup = bs(page.text, "html.parser")

    return soup


# rows = 2  # 2개의 사이트를 저장
# cols = 2  # 사이트의 json url의 이름과 api_key 값 저장(2개)
# arr = [[0 for i in range(cols)] for j in range(rows)]  # 예시 [url_name, api_key], [url_name, api_key], ....   ==> 나중에 여러 사이트를 돌릴때는 이렇게 하는게 좋을듯?


# python은 인터프리터 언어이기 때문에 순서를 잘 생각해야 한다.

arr = [['ozkiz1', 'KU5HdZg4BVXlfoLDEPu6EC'],['nsmall2022','KU5HdZg4BVXlfoLDEPu6EC'],['themedicube','0000']]

#일단 오즈킺즈의 3757과 농심의 3757이 겹치므로 두개만 해보겠다. + 메디큐브도 html이 파싱되는지 확인해보기
row = 2
for i in range(0, 3) :
    for k in range(3756, 3760) :
        list = arr[i][0]
        app_key=arr[i][1]

        requestdata = func1(list, k, app_key)
        
        if requestdata.status_code == 200 :
            jsonData = requestdata.json()
            for data in jsonData :
                if jsonData.get(data).get("price") != None :
                    print(f"{arr[i][0]}의 {k}페이지 입니다.----------------------------------------------")
                    ws[f'A{row}'] = jsonData.get(data).get("price")
                    ws[f'C{row}'] = jsonData.get(data).get("product_code")
                    ws[f'E{row}'] = jsonData.get(data).get("detail_image")

                    row = row + 1
                        
                    print(data, " : ", jsonData.get(data).get("price"), ", product_code", " : ", jsonData.get(data).get("product_code"), ", image_url : ", jsonData.get(data).get("detail_image"))
                else :
                    continue
        else : 
            print(requestdata)
        

wb.save(r'D:\WebCrawling\파이썬엑셀다루기\참가자_data.xlsx')