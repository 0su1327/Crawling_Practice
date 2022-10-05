import requests
import json
import openpyxl


wb = openpyxl.Workbook()

ws = wb.create_sheet('ozkids 크롤링')

ws['A1'] = 'product_price'
ws['B1'] = 'product_code'

row = 2
#  ozkiz1.com
#  discoveryglobal


# 일단 none 값을 저장 안하는 방식을 찾아봄(어차피 json이 없으면 아무값도 없는 거니까 아무값이나 하나 잡아서 none이 아닐때만 출력)
for i in range(5010,5030) :
    
    try:
        requestdata = requests.get(f"https://ozkiz1.cafe24api.com/api/v2/products/{i}?shop_no=1&cafe24_app_key=KU5HdZg4BVXlfoLDEPu6EC&cafe24_api_version=2022-06-01")

        # jsonData = None(default값)
        if requestdata.status_code == 200 :
            jsonData = requestdata.json()
            for data in jsonData :
                if jsonData.get(data).get("price") != None :
                    print(f"{i}페이지 입니다.---------------------------")
                    ws[f'A{row}'] = jsonData.get(data).get("price")
                    ws[f'B{row}'] = jsonData.get(data).get("product_code")

                    row = row + 1
                    
                    print(data, " : ", jsonData.get(data).get("price"), ", product_code", " : ", jsonData.get(data).get("product_code"))

                else :
                    continue

    except:
        print("HTML루 파싱")
wb.save(r'D:\WebCrawling\파이썬엑셀다루기\참가자_data.xlsx')