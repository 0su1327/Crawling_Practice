import requests
import json
import openpyxl


wb = openpyxl.Workbook()

ws = wb.create_sheet('ozkids 크롤링')

ws['A1'] = 'product_price'
ws['B1'] = 'product_code'

row =2
for i in range(5010,5020) :
    print(f"{i}페이지 입니다.---------------------------")
    requestdata = requests.get(f"https://ozkiz1.cafe24api.com/api/v2/products/{i}?shop_no=1&cafe24_app_key=KU5HdZg4BVXlfoLDEPu6EC&cafe24_api_version=2022-06-01")

    jsonData = None
    if requestdata.status_code == 200 :
        jsonData = requestdata.json()
        for data in jsonData :
            ws[f'A{row}'] = jsonData.get(data).get("price")
            ws[f'B{row}'] = jsonData.get(data).get("product_code")

            row = row + 1
            
            print(data, " : ", jsonData.get(data).get("price"), " : ", jsonData.get(data).get("product_code"))



wb.save(r'D:\WebCrawling\파이썬엑셀다루기\참가자_data.xlsx')