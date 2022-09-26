import requests

from bs4 import BeautifulSoup

import openpyxl

#엑셀파일 불러오기
wb = openpyxl.load_workbook(r'D:\WebCrawling\파이썬엑셀다루기\참가자_data.xlsx')
ws = wb.active #현재 활성화된 시트를 선택

#종목 코드 리스트

codes = [
    '005930',
    '000660',
    '035720'
]

row = 2 
for code in codes:
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    price = soup.select_one("#_nowVal").text
    high = soup.select_one('#_high').text
    price = price.replace(',','')
    high = high.replace(',','')
    print(price)
    print(high)
    ws[f'A{row}'] = '현재가'
    ws[f'D{row}'] = '고가'
    ws[f'B{row}'] = int(price) 
    ws[f'C{row}'] = int(high)
    row = row + 1  #B2 ,B3, B4 순으로 커진다.

wb.save(r'D:\WebCrawling\파이썬엑셀다루기\참가자_data.xlsx')