import openpyxl

fpath = r'D:\WebCrawling\파이썬엑셀다루기\참가자_data.xlsx'

# 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 엑셀 시트선택
ws = wb['오징어게임']

# 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '성기훈'

# 엑셀 저장하기
wb.save(fpath)